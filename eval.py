"""
evaluar.py — Offline retrieval evaluation against the organisers' sample
ground truth (CODEFEST AD ASTRA 2026). NOT part of entrega/.

The ground truth lives in `FASE ORDENADA CODEFEST.xlsx`: two sheets (F1, F3),
columns [id, PREGUNTA, FRAGMENTO, DOCUMENTO]. A question occupies one row and
its extra relevant fragments occupy the rows below it with the first two
columns blank, so question text is forward-filled down the sheet.

WHY MATCHING IS DONE BY TEXT, NOT BY chunk_id
    Some GT rows carry a chunk_id like DOC-0292-chunk-0144. Those come from
    one particular build and cannot be relied on: those numbers depend on
    rglob ordering and on max_chars/overlap_chars, so changing either makes
    every annotated id point at different text -- a metric collapse that
    looks like a retrieval regression and is not one.

    Since the FAQ, chunk_id is the FAISS index (0, 1, 2, ...), which is even
    more build-specific. That is the right choice for the SUBMISSION -- the
    organisers asked for it, and 10.2.1 judges fragments on `text` anyway --
    but it makes text matching the only sound basis for scoring against a
    ground truth annotated on someone else's build. Hence this function.

    Pooled judgments from tools/build_pool.py are the exception: they are
    produced by the index being scored, so there chunk_id matching is exact
    and correct.

WHY A GT FRAGMENT CAN MATCH TWO CHUNKS
    GT fragments are ~1050 chars and the index chunks at ~1000 with ~350 of
    overlap, so a GT fragment routinely straddles a chunk boundary. Credit
    goes to any chunk covering at least --frag-threshold of the fragment's
    n-grams, and each GT fragment is credited at most ONCE.

READ THE CEILING BEFORE READING THE SCORE
    F1@3 cannot reach 1.0 when a query has fewer than 3 relevant documents.
    With |D*| = 1 the best possible is P=1/3, R=1, F1=0.5. The summary prints
    the per-query ceiling next to your score, because "0.500 on q003" is a
    perfect result and "0.500 on q002" is a third of one, and a mean that
    mixes them tells you nothing on its own.

Usage:
    python evaluar.py --gt "FASE ORDENADA CODEFEST.xlsx" \
        --index-dir entrega/base_vectorial

    # look at what the pipeline actually returned, with the text
    python evaluar.py --gt ... --show-fragments 10

    # is the GT text even in the index?
    python evaluar.py --gt ... --audit
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import unicodedata
from pathlib import Path


from entrega.generador import (GraphIndex, LexicalIndex, RetrievalConfig,   # noqa: E402
                       VectorStore, add_retrieval_args, aggregate_documents,
                       aggregate_documents_rankdecay, config_from_args,
                       load_stores, read_jsonl, retrieve)

NGRAM = 8                 # word n-gram size for fragment matching
ID_NGRAM = 5              # smaller n for matching GT questions to official ids


# ------------------------------------------------- normalisation

def norm_doc(name: str) -> str:
    """
    Collapse the GT naming convention onto the corpus one.

    THE GRADERS MATCH ON doc_id, NOT ON fuente. Confirmed in the organisers'
    FAQ, three times:

        "La última frase del cuadro de la Sección 10.2 corresponde a un error
         de versionamiento... El emparejamiento a nivel de documento se
         realiza mediante el campo doc_id, el cual será suministrado por ADL."
        "Deben emparejar los resultados mediante ese DOC_ID oficial y
         conservar fuente como metadata."

    So `doc_id` in every record MUST be the DOC_ID from
    Indice_Datos_Codefest.xlsx (F1-AIINDEX-001 and so on), which is what
    inventory.py already supplies -- and the whole `fuente` path-vs-basename
    question turns out not to have mattered.

    This function survives only because the SAMPLE ground truth in the xlsx
    identifies documents by FILE NAME. To score locally the name still has to
    be resolved to a doc_id, via inventory.json. The real evaluation will not
    need it.

    GT:     daio_study2529_guarding_the_alliances_..._dziegiel.pdf
    disk:   DAIO_study2529-guarding-the-alliances-..._dziegiel.pdf
    """
    name = Path(name.strip()).name.lower()
    return re.sub(r"[_\-\s]+", "", name)


def norm_text(text: str) -> str:
    text = unicodedata.normalize("NFC", text).lower()
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", text)).strip()


def ngrams(text: str, n: int = NGRAM) -> set[str]:
    words = norm_text(text).split()
    if len(words) < n:
        return {" ".join(words)} if words else set()
    return {" ".join(words[i:i + n]) for i in range(len(words) - n + 1)}


def coverage(gt_fragment: set[str], chunk: set[str]) -> float:
    """Share of the GT fragment's n-grams present in the chunk."""
    if not gt_fragment:
        return 0.0
    return len(gt_fragment & chunk) / len(gt_fragment)


# ------------------------------------------------- ground truth

def load_ground_truth(path: Path) -> list[dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    queries: list[dict] = []

    for sheet in workbook.worksheets:
        current: dict | None = None
        for row in sheet.iter_rows(values_only=True):
            cells = [(str(c).strip() if c is not None else "") for c in row]
            cells += [""] * (4 - len(cells))
            _id, question, fragment, document = cells[:4]

            if question.upper() == "PREGUNTA":
                continue
            if question:
                current = {"sheet": sheet.title, "gt_id": _id, "query": question,
                           "fragments": [], "documents": []}
                queries.append(current)
            if current is None or not fragment:
                continue

            # DOCUMENTO holds "file.pdf" or "file.pdf\nDOC-xxxx-chunk-yyyy"
            filename = next((l.strip() for l in document.splitlines()
                             if l.strip() and not l.strip().startswith("DOC-")), "")
            current["fragments"].append({"text": fragment, "document": filename})
            if filename and filename not in current["documents"]:
                current["documents"].append(filename)

    return [q for q in queries if q["fragments"]]



def load_judgments(path: Path) -> list[dict]:
    """
    Read pooled relevance judgments produced by tools/build_pool.py.

    Returns the same shape load_ground_truth() does, so evaluate() needs no
    special case -- except that relevance is GRADED (0/1/2) and keyed by
    chunk_id, which is exact. The xlsx sample GT matches on text overlap
    because its chunk_ids came from a different build; these were produced by
    the index being scored, so no fuzzy matching is needed or wanted.
    """
    import openpyxl

    sheet = openpyxl.load_workbook(path, data_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c or "").strip().lower() for c in rows[0]]

    def column(name: str) -> int:
        return header.index(name) if name in header else -1

    idx = {name: column(name) for name in
           ("query_id", "consulta", "relevancia", "chunk_id", "doc_id")}
    if idx["query_id"] < 0 or idx["relevancia"] < 0:
        raise SystemExit(f"{path} has no query_id/relevancia columns: {header}")

    by_query: dict[str, dict] = {}
    judged = graded = 0
    for row in rows[1:]:
        query_id = str(row[idx["query_id"]] or "").strip()
        if not query_id:
            continue
        raw = row[idx["relevancia"]]
        # Blank means "judged not relevant". A pooled fragment nobody marked
        # is a fragment a human looked at and passed over.
        try:
            grade = int(float(raw)) if raw not in (None, "") else 0
        except (TypeError, ValueError):
            grade = 0
        judged += 1
        graded += grade > 0

        entry = by_query.setdefault(query_id, {
            "query_id": query_id, "gt_id": query_id, "sheet": "pool",
            "query": str(row[idx["consulta"]] or ""),
            "fragments": [], "documents": [], "grades": {}})
        chunk_id = str(row[idx["chunk_id"]] or "")
        doc_id = str(row[idx["doc_id"]] or "")
        entry["grades"][chunk_id] = grade
        if grade > 0:
            entry["fragments"].append({"text": "", "document": doc_id,
                                       "chunk_id": chunk_id})
            if doc_id and doc_id not in entry["documents"]:
                entry["documents"].append(doc_id)

    print(f"  judgments: {judged} rows, {graded} relevant, "
          f"{len(by_query)} queries")
    unjudged = [q for q, e in by_query.items() if not e["fragments"]]
    if unjudged:
        print(f"  {len(unjudged)} queries have NO relevant fragment marked: "
              f"{unjudged[:6]}\n    Those score 0 by definition. Either they "
              f"are genuinely unanswerable from\n    this corpus, or the pool "
              f"missed the answer -- worth checking before trusting the mean.")
    return [e for e in by_query.values() if e["fragments"]]


def resolve_ids(gt: list[dict], queries_path: Path) -> None:
    """
    Attach the official q001..q050 id to each GT question by text match.

    The sheet numbers its questions 2/3/4 and q0047..q0052, which are NOT the
    official ids: they come from an earlier revision of the query set. Seven
    match the official text (q002-q004, q033-q036). Two do not exist in the
    final 50 at all -- both about water security in Central America -- and
    scoring against those inflates the average with questions never asked.
    """
    official = read_jsonl(queries_path)

    for entry in gt:
        grams = ngrams(entry["query"], ID_NGRAM)
        score, best = max(
            ((len(grams & ngrams(o["query"], ID_NGRAM)) / max(len(grams), 1),
              o["query_id"]) for o in official), default=(0.0, None))
        entry["query_id"] = best if score > 0.8 else None


# ------------------------------------------------- metrics

def f1_at_k(hits: list[bool], n_relevant: int, k: int) -> float:
    """
    Spec 10.2.2: a SET metric, order-insensitive. Recall's denominator is
    min(|D*|, k) so a query with fewer than k relevant documents is not
    penalised for slots it cannot possibly fill.
    """
    found = sum(hits[:k])
    precision = found / k
    recall = found / min(n_relevant, k) if n_relevant else 0.0
    if precision + recall == 0:
        return 0.0
    return 2 * precision * recall / (precision + recall)


def f1_ceiling(n_relevant: int, k: int = 3) -> float:
    """Best achievable F1@k for a query with n_relevant documents."""
    return f1_at_k([True] * min(n_relevant, k), n_relevant, k)


def ndcg(hits: list[bool], k: int, n_relevant: int) -> float:
    """Binary-gain nDCG@k. The ideal ranking puts every relevant item on top."""
    dcg = sum(1.0 / math.log2(rank + 1)
              for rank, hit in enumerate(hits[:k], 1) if hit)
    ideal = sum(1.0 / math.log2(rank + 1)
                for rank in range(1, min(n_relevant, k) + 1))
    return dcg / ideal if ideal else 0.0


def reciprocal_rank(hits: list[bool]) -> float:
    for rank, hit in enumerate(hits, 1):
        if hit:
            return 1.0 / rank
    return 0.0


# ------------------------------------------------- evaluation

def _aggregate(doc_ranking, n, pool, cfg: RetrievalConfig):
    """
    FIX: mirrors generador.py main()'s dispatch on cfg.doc_score exactly.

    Before this fix, evaluate() called aggregate_documents() (the old
    cosine/CombSUM aggregator) unconditionally, regardless of cfg.doc_score.
    That was harmless when doc_score defaulted to "cosine", but the default
    is now "rankdecay" (see generador.py's RetrievalConfig docstring), under
    which retrieve() returns doc_ranking as raw RRF-space scores rather than
    cosine/CombSUM ones. Feeding RRF-space scores (~0.003-0.017) into
    aggregate_documents()'s max-pooling and hit_bonus -- both calibrated
    against a cosine spread of ~0.11 or a CombSUM spread of ~3.0 -- silently
    produces a document ranking that has no relationship to the one
    resultados.jsonl actually ships, which is exactly the "measuring a
    different system than you ship" failure this module's own docstring
    warns about. Every call site that aggregates result.doc_ranking must go
    through this function, not aggregate_documents() directly.
    """
    if cfg.doc_score == "rankdecay":
        return aggregate_documents_rankdecay(doc_ranking, n, pool, cfg.doc_decay)
    return aggregate_documents(doc_ranking, n, pool, cfg.doc_hit_bonus,
                               cfg.doc_hit_cap, cfg.doc_agg, cfg.doc_top_m)


def evaluate(gt, stores, cfg: RetrievalConfig, lexical, graph, frag_threshold: float,
             doc_k: int, frag_k: int, show_misses: bool,
             show_fragments: int, show_docs: bool) -> None:
    """
    Score the SHIPPED pipeline, not a simplified stand-in.

    Ranking goes through generador.retrieve() and the CLI is built from
    generador.add_retrieval_args(), so there is exactly one place where a
    default can be changed and it changes both sides at once. Document
    aggregation goes through _aggregate(), for the same reason -- see its
    docstring.
    """
    rows = []

    for entry in gt:
        result = retrieve(stores, entry["query"], entry.get("query_id", ""),
                          cfg, lexical, graph)

        # ---- document level: aggregate over the same list main() uses
        wanted = [norm_doc(d) for d in entry["documents"] if d]
        ranked_docs = _aggregate(result.doc_ranking, doc_k, cfg.doc_pool, cfg)
        by_doc_id = {m["doc_id"]: m for m, _ in result.doc_ranking}

        def matches(doc_id: str) -> bool:
            if entry.get("grades") is not None:
                return doc_id in entry["documents"]   # official ids, exact
            meta = by_doc_id.get(doc_id, {})
            source = norm_doc(meta.get("nombre_archivo") or meta.get("fuente", ""))
            return any(source.endswith(w) or w.endswith(source)
                       for w in wanted if w)

        doc_hits = [matches(d) for d in ranked_docs]

        # WHERE the GT documents landed in the FULL ranking, not just whether
        # they cleared rank 3. This is the number that moves when you change a
        # flag; F1@3 only moves when a document crosses the top-3 cliff, which
        # on seven queries makes most real improvements invisible.
        # DEEP pool on purpose. Aggregating at cfg.doc_pool would truncate the
        # ranking to the ~15 documents represented in the top 30 chunks, and a
        # GT document just outside that would report as [] -- indistinguishable
        # from "never retrieved". The shipped top-3 still uses cfg.doc_pool;
        # this is only the ruler.
        full_docs = _aggregate(result.doc_ranking, None,
                               len(result.doc_ranking), cfg)
        doc_ranks = [i for i, d in enumerate(full_docs, 1) if matches(d)]

        # Rank under the SHIPPED pool too, so the gap between "the ranker
        # knows" and "the aggregation kept it" is visible.
        pooled = _aggregate(result.doc_ranking, None, cfg.doc_pool, cfg)
        pooled_ranks = [i for i, d in enumerate(pooled, 1) if matches(d)]

        # ---- fragment level
        grades = entry.get("grades")
        if grades is not None:
            # Pooled judgments: exact chunk_id match, graded relevance.
            frag_hits = [grades.get(m["chunk_id"], 0) > 0
                         for m, _s in result.unique[:frag_k]]
            n_relevant = sum(1 for g in grades.values() if g > 0)
            row = {
                "doc_ranks": doc_ranks, "frag_ranks": [],
                "pool": len(result.candidates),
                "label": entry["query_id"], "query": entry["query"],
                "n_docs": len(wanted),
                "f1": f1_at_k(doc_hits, len(wanted), doc_k),
                "f1_max": f1_ceiling(len(wanted), doc_k),
                "ndcg": ndcg(frag_hits, frag_k, n_relevant),
                "mrr": reciprocal_rank(frag_hits),
                "covered": sum(frag_hits), "n_frags": n_relevant,
            }
            rows.append(row)
            print(f"  {row['label']:12s} F1@{doc_k}={row['f1']:.3f}/"
                  f"{row['f1_max']:.3f}  nDCG@{frag_k}={row['ndcg']:.3f}  "
                  f"hits {row['covered']}/{min(n_relevant, frag_k)}")
            continue

        gt_grams = [ngrams(f["text"]) for f in entry["fragments"]]
        frag_hits, covered = [], set()
        for meta, _score in result.unique[:frag_k]:
            chunk_grams = ngrams(meta["texto"])
            # Credit each GT fragment ONCE. A fragment spanning a chunk
            # boundary matches two chunks, and counting both pushes DCG above
            # IDCG -- which is how nDCG > 1 happens.
            fresh = {i for i, grams in enumerate(gt_grams)
                     if i not in covered
                     and coverage(grams, chunk_grams) >= frag_threshold}
            frag_hits.append(bool(fresh))
            covered |= fresh

        # Deepest position in the WHOLE candidate pool at which each GT
        # fragment is covered. "not in the top 10" and "at position 800" call
        # for completely different work, and F1/nDCG cannot tell them apart.
        frag_ranks = []
        for grams in gt_grams:
            hit = next((r for r, (m, _s) in enumerate(result.candidates, 1)
                        if coverage(grams, ngrams(m["texto"])) >= frag_threshold),
                       None)
            frag_ranks.append(hit)

        row = {
            "doc_ranks": doc_ranks,
            "frag_ranks": frag_ranks,
            "pool": len(result.candidates),
            "label": entry.get("query_id") or f"{entry['sheet']}/{entry['gt_id']}",
            "query": entry["query"],
            "n_docs": len(wanted),
            "f1": f1_at_k(doc_hits, len(wanted), doc_k),
            "f1_max": f1_ceiling(len(wanted), doc_k),
            "ndcg": ndcg(frag_hits, frag_k, len(gt_grams)),
            "mrr": reciprocal_rank(frag_hits),
            "covered": len(covered),
            "n_frags": len(gt_grams),
        }
        rows.append(row)

        def fmt(values, total):
            return "[" + ",".join(str(v) if v else f">{total}" for v in values) + "]"

        print(f"  {row['label']:12s} F1@{doc_k}={row['f1']:.3f}/{row['f1_max']:.3f}  "
              f"nDCG@{frag_k}={row['ndcg']:.3f}  "
              f"gt_doc_rank={fmt(pooled_ranks, cfg.doc_pool)}"
              f"/deep{fmt(doc_ranks, len(full_docs))} "
              f"gt_frag_rank={fmt(frag_ranks, row['pool'])}")
        print(f"               {entry['query'][:70]}...")

        if show_docs:
            print(f"      GT docs : {[d for d in entry['documents'] if d]}")
            for rank, doc_id in enumerate(ranked_docs, 1):
                meta = by_doc_id.get(doc_id, {})
                mark = "HIT " if doc_hits[rank - 1] else "    "
                print(f"      {mark}{rank}. {doc_id}  "
                      f"{Path(meta.get('fuente', '?')).name[:64]}")

        if show_fragments:
            inspect_fragments(result, entry, gt_grams, frag_threshold,
                              show_fragments)

        if show_misses and len(covered) < len(gt_grams):
            for i, fragment in enumerate(entry["fragments"]):
                if i in covered:
                    continue
                # key= is required: on a tie in coverage (very common when
                # every chunk scores 0.00) max() falls through to comparing
                # metadata dicts, which are not orderable.
                best = max(((coverage(gt_grams[i], ngrams(m["texto"])), m)
                            for m, _ in result.unique[:frag_k]),
                           key=lambda pair: pair[0], default=(0.0, None))
                print(f"      MISS  doc={fragment['document'][:44]}  "
                      f"best_coverage={best[0]:.2f}")
                print(f"            gt: {norm_text(fragment['text'])[:90]}")

    summarise(rows, doc_k, frag_k)


def inspect_fragments(result, entry, gt_grams, frag_threshold: float,
                      show: int) -> None:
    """
    Print the top fragments with their text so they can be judged by eye.

    This is the diagnostic that every score-only report is missing. A zero
    nDCG has three completely different causes and they need opposite fixes:

      PLAUSIBLE    the winners are on-topic prose from sensible documents and
                   the GT chunk is a hair behind. That is the encoder's
                   resolution limit; only a stronger ranking signal (lexical
                   channel, cross-encoder) moves it.
      REDUNDANT    several slots hold near-identical text from one document.
                   Retrieval is fine, the slots are wasted --
                   lower --dedupe-threshold. Cheap, no rebuild.
      BOILERPLATE  the winners are tables of contents, running heads,
                   reference lists, acknowledgements: text that is close to
                   everything and specific to nothing. That is an
                   extraction/indexing problem and no ranking change fixes it.

    The `channels` column shows each fragment's rank in each retrieval
    channel, so you can see whether BM25 and the dense encoders agree or
    whether one of them is carrying the result alone.
    """
    ranks_by_channel = []
    for name, ranking in result.channels:
        ranks_by_channel.append(
            (name.split("/")[-1][:14],
             {meta["chunk_id"]: r for r, (meta, _s) in enumerate(ranking, 1)}))

    print(f"      {'-' * 92}")
    print(f"      TOP {show} FRAGMENTS RETURNED"
          f"   (channels: {', '.join(n for n, _ in ranks_by_channel)})")

    seen_shingle_sets: list[set] = []
    for rank, (meta, score) in enumerate(result.unique[:show], 1):
        chunk_grams = ngrams(meta["texto"])
        best_cover = max((coverage(g, chunk_grams) for g in gt_grams), default=0.0)

        positions = " ".join(
            f"{name}={lookup.get(meta['chunk_id'], '-')}"
            for name, lookup in ranks_by_channel)

        mine = set(chunk_grams)
        near = any(theirs and len(mine & theirs) / max(min(len(mine), len(theirs)), 1) >= 0.3
                   for theirs in seen_shingle_sets)
        seen_shingle_sets.append(mine)

        flag = "<== GT" if best_cover >= frag_threshold else ("~dup" if near else "")
        print(f"      {rank:2d}. s={score:.4f} cov={best_cover:.2f} "
              f"F{meta.get('fenomeno', '?')} p{meta.get('posicion', 0):<4d} "
              f"{Path(meta.get('fuente', '?')).name[:44]:44s} {flag}")
        print(f"          [{positions}]")
        if meta.get("contexto"):
            print(f"          ctx: {meta['contexto'][:100]}")
        print(f"          {norm_text(meta['texto'])[:180]}")

    docs = {m["doc_id"] for m, _ in result.unique[:show]}
    print(f"      distinct documents in top {show}: {len(docs)}/{show}")
    print(f"      {'-' * 92}")


def summarise(rows: list[dict], doc_k: int, frag_k: int) -> None:
    n = len(rows) or 1
    print()
    print(f"  {'query':10s} {'|D*|':>5s} {'F1@3 max':>9s} {'F1@3':>7s} "
          f"{'nDCG@10':>8s} {'frags':>7s}")
    print("  " + "-" * 52)
    for row in rows:
        mark = " *" if row["f1"] >= row["f1_max"] - 1e-9 and row["f1"] > 0 else ""
        print(f"  {row['label']:10s} {row['n_docs']:5d} {row['f1_max']:9.3f} "
              f"{row['f1']:7.3f} {row['ndcg']:8.3f} "
              f"{row['covered']:3d}/{row['n_frags']:<3d}{mark}")
    print("  " + "-" * 52)
    print(f"  {'MEAN':10s} {'':5s} {sum(r['f1_max'] for r in rows)/n:9.3f} "
          f"{sum(r['f1'] for r in rows)/n:7.3f} "
          f"{sum(r['ndcg'] for r in rows)/n:8.3f}")

    # Continuous companions to the two cliff metrics. Mean reciprocal rank
    # over the FULL rankings responds to a document moving 40 -> 5, which
    # F1@3 cannot see. Tune against these; report the official ones.
    def mrr(key):
        total = 0.0
        for row in rows:
            best = min((v for v in row[key] if v), default=None)
            total += 1.0 / best if best else 0.0
        return total / n

    found_docs = sum(1 for r in rows if any(r["doc_ranks"]))
    found_frags = sum(1 for r in rows if any(v for v in r["frag_ranks"]))
    print(f"\n  --- continuous diagnostics (not the official metrics) ---")
    print(f"  MRR over the full document ranking : {mrr('doc_ranks'):.4f}"
          f"   ({found_docs}/{n} queries have a GT doc anywhere)")
    print(f"  MRR over the full fragment pool    : {mrr('frag_ranks'):.4f}"
          f"   ({found_frags}/{n} queries have a GT fragment anywhere)")
    print(f"  These move when a flag helps even if F1@3 does not. If a GT doc "
          f"is\n  >pool it never entered the candidate list at all -- raise "
          f"--depth or\n  --doc-pool before touching anything else.")

    print(f"\n  F1@{doc_k}    documents (10.2.2): {sum(r['f1'] for r in rows)/n:.4f}"
          f"   ceiling {sum(r['f1_max'] for r in rows)/n:.4f}")
    print(f"  nDCG@{frag_k} fragments (10.2.1): {sum(r['ndcg'] for r in rows)/n:.4f}")
    print(f"  MRR      fragments (diagnostic): {sum(r['mrr'] for r in rows)/n:.4f}")
    print(f"\n  {n} queries is far too few for a stable estimate, and the real "
          "50-query ground truth\n  has its own |D*| distribution. Use these "
          "numbers to catch regressions, not to tune\n  decimals. A change "
          "worth under ~0.05 here is noise.")


# ------------------------------------------------- audit

def audit(gt: list[dict], store: VectorStore, threshold: float) -> None:
    """
    Is the GT text in the index AT ALL, ignoring the ranker?

    A zero score has two opposite causes: the text never made it into the
    index (extraction, OCR, or a missing file), or it is indexed but ranked
    below the cut. The fix for one is useless against the other, so establish
    which it is before changing anything about the encoder or the chunking.
    """
    lookup: dict[str, set[int]] = {}
    flat = [(qi, f) for qi, q in enumerate(gt) for f in q["fragments"]]
    for k, (_qi, fragment) in enumerate(flat):
        for gram in ngrams(fragment["text"]):
            lookup.setdefault(gram, set()).add(k)

    sizes = [len(ngrams(f["text"])) for _, f in flat]
    best = [(0.0, None) for _ in flat]

    print(f"\n  Scanning {len(store.metadata)} chunks for {len(flat)} GT fragments ...")
    for meta in store.metadata:
        counts: dict[int, int] = {}
        for gram in ngrams(meta["texto"]):
            for k in lookup.get(gram, ()):
                counts[k] = counts.get(k, 0) + 1
        for k, count in counts.items():
            score = count / max(sizes[k], 1)
            if score > best[k][0]:
                best[k] = (score, meta)

    print(f"\n  {'query':8s} {'cover':>6s}  {'verdict':22s} chunk source")
    absent = 0
    for k, (qi, fragment) in enumerate(flat):
        score, meta = best[k]
        if score >= threshold:
            verdict = "indexed, RANKING issue"
        elif score >= 0.10:
            verdict = "partial -- CHUNKING?"
        else:
            verdict = "ABSENT from index"
            absent += 1
        source = Path(meta["fuente"]).name[:46] if meta else "-"
        print(f"  {gt[qi].get('query_id') or '?':8s} {score:6.2f}  {verdict:22s} {source}")
        if score < 0.10:
            print(f"           GT doc: {fragment['document'][:60]}")

    print(f"\n  {absent}/{len(flat)} GT fragments are not in the index at all.")
    if absent:
        print("  Those are an extraction/corpus problem, not a retrieval one.")


_OFFICIAL_DOC_ID = re.compile(r"^F[123]-[A-Z0-9]+-\d+$")


def check_doc_ids(store: VectorStore) -> None:
    """
    The graders match on doc_id, so every doc_id must be the OFFICIAL one from
    Indice_Datos_Codefest.xlsx. A locally-generated id like DOC-00292 matches
    nothing and scores zero with no error -- the same silent failure the
    `fuente` question threatened, now relocated to the field that actually
    counts.
    """
    ids = {m.get("doc_id", "") for m in store.metadata}
    official = {d for d in ids if _OFFICIAL_DOC_ID.match(d)}
    print(f"\n  doc_id: {len(ids)} distinct, {len(official)} match the "
          f"official F#-OBS-### pattern")
    print(f"  sample: {sorted(ids)[:3]}")

    if len(official) < len(ids):
        stray = sorted(ids - official)[:5]
        print(f"  WARNING: {len(ids) - len(official)} doc_ids are NOT official "
              f"ADL ids, e.g. {stray}.\n"
              f"           10.2.2 matching is by doc_id (FAQ). Those "
              f"documents cannot score.\n"
              f"           They are files inventory.py failed to match "
              f"against Indice_Datos_Codefest.xlsx.")


# ------------------------------------------------- entry point

def main() -> None:
    parser = argparse.ArgumentParser()
    add_retrieval_args(parser)
    parser.set_defaults(index_dir=Path("entrega/base_vectorial"))
    parser.add_argument("--gt", type=Path, default=None,
                        help="FASE ORDENADA CODEFEST.xlsx (7-query sample)")
    parser.add_argument("--judgments", type=Path, default=None,
                        help="pool.xlsx from tools/build_pool.py, judged by "
                             "hand. 50 queries instead of 7, graded relevance, "
                             "exact chunk_id matching.")
    parser.add_argument("--queries", type=Path, default=Path("consultas_50.jsonl"))
    parser.add_argument("--doc-k", type=int, default=3)
    parser.add_argument("--frag-k", type=int, default=10)
    parser.add_argument("--frag-threshold", type=float, default=0.30,
                        help="share of the GT fragment's n-grams a chunk must "
                             "contain to count as relevant")
    parser.add_argument("--keep-unmapped", action="store_true",
                        help="also score GT questions absent from the final 50")
    parser.add_argument("--dump-gt", type=Path, default=None)
    parser.add_argument("--show-misses", action="store_true")
    parser.add_argument("--show-fragments", type=int, default=0, metavar="N",
                        help="print the top N returned fragments WITH THEIR "
                             "TEXT and per-channel ranks, for manual "
                             "inspection. This is the diagnostic that tells "
                             "you whether a zero is a ranking problem, a "
                             "redundancy problem or an extraction problem.")
    parser.add_argument("--show-docs", action="store_true",
                        help="print the 3 returned documents next to the GT ones")
    parser.add_argument("--audit", action="store_true")
    args = parser.parse_args()
    cfg = config_from_args(args)

    if not args.gt and not args.judgments:
        raise SystemExit("Pass --gt (the 7-query sample) or --judgments "
                         "(your pooled judgments).")

    if args.judgments:
        gt = load_judgments(args.judgments)
        args.queries = Path("/nonexistent")     # ids are already official
    else:
        gt = load_ground_truth(args.gt)
    if args.queries.exists():
        resolve_ids(gt, args.queries)
        unmapped = [q for q in gt if not q.get("query_id")]
        if unmapped and not args.keep_unmapped:
            for q in unmapped:
                print(f"  dropping {q['sheet']}/{q['gt_id']}: not in the final "
                      f"50 -- {q['query'][:60]}")
            gt = [q for q in gt if q.get("query_id")]
        gt.sort(key=lambda q: q.get("query_id") or "")
    else:
        print(f"  {args.queries} not found -- ids not resolved")

    print(f"Ground truth: {len(gt)} queries, "
          f"{sum(len(q['fragments']) for q in gt)} fragments")

    if args.dump_gt:
        with args.dump_gt.open("w", encoding="utf-8") as fh:
            for entry in gt:
                fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
        print(f"Wrote {args.dump_gt}")

    stores = load_stores(args.index_dir, args.doc_encoder, cfg.doc_score)
    print(f"evaluar  : build 2026-08-08d (reports gt_doc_rank / gt_frag_rank)")
    print(f"Encoders : {[s.name for s in stores]}")
    print(f"Documents: {cfg.doc_score}"
          + (f", decay={cfg.doc_decay}" if cfg.doc_score == "rankdecay"
             else f", pool={cfg.doc_pool}, hit +{cfg.doc_hit_bonus:.0%} "
                  f"x{cfg.doc_hit_cap}"))
    print(f"Fragments: dedupe>={cfg.dedupe_threshold}, "
          f"phenomenon {cfg.phenomenon_mode} {cfg.phenomenon_boost}, "
          f"bm25 w={cfg.bm25_weight}"
          + (f", reranker={cfg.reranker}" if cfg.reranker else ""))
    check_doc_ids(stores[0])

    lexical = None
    if cfg.bm25_weight > 0:
        print(f"Lexical  : building BM25 over {len(stores[0].metadata)} chunks ...")
        lexical = LexicalIndex.build(stores[0].metadata,
                                     [q["query"] for q in gt],
                                     cfg.bm25_k1, cfg.bm25_b)

    graph = None
    if cfg.graph_weight > 0:
        graph = GraphIndex.load(args.index_dir / "grafo" / "grafo.graphml")
        if graph is not None:
            print(f"Grafo    : {len(graph.entities)} entities, "
                  f"{sum(len(v) for v in graph.neighbours.values()) // 2} "
                  f"relations")
        else:
            print(f"Grafo    : none at "
                  f"{args.index_dir / 'grafo' / 'grafo.graphml'} -- "
                  f"bonus not built, or --graph-weight 0 to silence this")
    print()

    evaluate(gt, stores, cfg, lexical, graph, args.frag_threshold, args.doc_k,
             args.frag_k, args.show_misses, args.show_fragments, args.show_docs)

    if args.audit:
        audit(gt, stores[0], args.frag_threshold)


if __name__ == "__main__":
    main()